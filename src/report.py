"""Reporting: Excel workbook, figures, and terminal Markdown — target-agnostic."""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

import pandas as pd

logger = logging.getLogger("docksuite.report")


class ReportGenerator:
    """Emit Excel, figures, and a Markdown summary for a benchmark/screen.

    Args:
        output_dir: Directory for the workbook and figures.
        generate_figures: Whether to render matplotlib figures.
    """

    def __init__(self, output_dir: Path, generate_figures: bool = True) -> None:
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.generate_figures = generate_figures

    def generate(self, rankings: pd.DataFrame, metrics, config,
                 control: Optional[dict] = None, extra_sheets: Optional[dict] = None) -> Path:
        """Write the report bundle and return the Excel path.

        Args:
            rankings: Ranked compounds (needs ``score``/``dG`` + ``is_active`` if labelled).
            metrics: A :class:`~vspipeline.metrics.BenchmarkMetrics` or ``None``.
            config: The active :class:`~vspipeline.config.PipelineConfig`.
            control: Optional redocking-control dict (rmsd, baseline dG).
            extra_sheets: Optional ``{sheet_name: DataFrame}`` additions.

        Returns:
            Path to the written ``.xlsx``.
        """
        pdb_id = config.get("target.pdb_id")
        xlsx = self.output_dir / f"{pdb_id}_benchmark_results.xlsx"
        with pd.ExcelWriter(xlsx, engine="openpyxl") as xw:
            self._sheet(xw, "Ranking", rankings)
            if metrics is not None:
                self._sheet(xw, "Metrics_Summary", pd.DataFrame([metrics.as_row()]))
            if control:
                self._sheet(xw, "Control_Validation", pd.DataFrame([control]))
            for name, frame in (extra_sheets or {}).items():
                self._sheet(xw, name[:31], frame)
        logger.info("wrote %s", xlsx.name)

        if self.generate_figures and metrics is not None and "is_active" in rankings:
            self._roc_figure(rankings, metrics, pdb_id)
        self._markdown_summary(rankings, metrics, config, control)
        return xlsx

    # ------------------------------------------------------------------ #
    @staticmethod
    def _sheet(writer, name: str, frame: pd.DataFrame) -> None:
        frame.to_excel(writer, sheet_name=name, index=False)
        ws = writer.sheets[name]
        try:
            from openpyxl.styles import Font, PatternFill
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.font = Font(bold=True, color="FFFFFF")
            ws.freeze_panes = "A2"
        except Exception:  # pragma: no cover
            pass

    def _roc_figure(self, rankings: pd.DataFrame, metrics, pdb_id: str) -> None:
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt
            from sklearn.metrics import roc_curve
        except Exception:  # pragma: no cover
            return
        score = rankings["score"] if "score" in rankings else -rankings["dG"]
        fpr, tpr, _ = roc_curve(rankings["is_active"], score)
        fig, ax = plt.subplots(figsize=(6, 6), dpi=300)
        ax.plot(fpr, tpr, lw=2, color="#d62728", label=f"AUC={metrics.roc_auc:.3f}")
        ax.plot([0, 1], [0, 1], ":", color="k", alpha=.4, label="random")
        ax.set_xlabel("FPR"); ax.set_ylabel("TPR"); ax.set_title(f"{pdb_id} ROC")
        ax.legend(loc="lower right")
        fig.tight_layout()
        path = self.output_dir / f"{pdb_id}_roc.png"
        fig.savefig(path); plt.close(fig)
        logger.info("wrote %s", path.name)

    @staticmethod
    def _markdown_summary(rankings: pd.DataFrame, metrics, config, control) -> None:
        lines = ["", "=" * 64, f"  {config.get('target.pdb_id')} BENCHMARK SUMMARY", "=" * 64]
        sort_col = "score" if "score" in rankings else "dG"
        top = rankings.sort_values(sort_col, ascending=("score" not in rankings)).head(15)
        lines.append("  Rank | Compound | score")
        for i, (_, r) in enumerate(top.iterrows(), 1):
            star = " *" if r.get("is_active") else ""
            val = r.get("score", r.get("dG"))
            lines.append(f"  {i:>2} | {r.get('id', r.name)}{star} | {val:.2f}")
        if metrics is not None:
            lines.append(f"\n  ROC-AUC : {metrics.roc_auc:.3f} "
                         f"CI95 [{metrics.auc_ci95[0]:.3f}, {metrics.auc_ci95[1]:.3f}]")
            lines.append(f"  EF      : " + " / ".join(f"{int(k*100)}%={v:.2f}"
                                                      for k, v in metrics.ef.items()))
            lines.append(f"  BEDROC  : {metrics.bedroc:.3f}")
        if control:
            lines.append(f"  Control : RMSD {control.get('redock_RMSD_A')} A, "
                         f"dG {control.get('baseline_dG')} kcal/mol")
        lines.append("=" * 64)
        logger.info("\n".join(lines))
